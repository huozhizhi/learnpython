import openpyxl
wb = openpyxl.load_workbook("/Users/xin.wang/Desktop/pytest.xlsx")
sht = wb['test']

for i in range(2,13):
	if sht.cell(i,1).value == sht.cell(i-1,1).value:
		sht.cell(i,2).value=sht.cell(i-1,2).value+1	
	else:
		sht.cell(i,2).value=1
	print(sht.cell(i,2).value)
	i=i+1
wb.save("/Users/xin.wang/Desktop/pytest.xlsx")